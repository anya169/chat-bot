from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Font
from .models import *
from pathlib import Path

def generate_report(cur_curator_login, ids):
   try:
      #получаем куратора и сотрудников
      curator = Employee.objects.filter(login = cur_curator_login).first()
      filtered_employees = Employee.objects.filter(id__in = ids)
      
      #загружаем шаблон
      current_dir = Path(__file__).parent
      template_path = current_dir / 'report_template.xlsx'     
      wb = load_workbook(template_path)
      ws = wb.active

      #получаем текущее время и форматируем его
      current_datetime = datetime.now().strftime("%d.%m.%Y %H:%M")

      #получаем все существующие опросы
      polls = Poll.objects.all()

      #заполняем шапку
      ws["A1"] = ws["A1"].value.replace("{{ date }}", current_datetime)
      ws["A2"] = ws["A2"].value.replace("{{ curator_name }}", curator.name)
      ws["A3"] = ws["A3"].value.replace("{{ count }}", str(filtered_employees.count()))
      
      #начинаем выводить информацию с 5 строки
      current_row = 5
      
      for poll in polls:   
         #сотрудники, прошедшие этот опрос
         current_employees = filtered_employees.filter(
            answer__question__poll=poll
         ).distinct()
        
         #если нет сотрудников, прошедших этот опрос, пропускаем его
         if not current_employees:
            continue
         #шапка таблицы 
         ws.cell(row=current_row, column=1, value=f"Опрос: {poll.name}").font = Font(bold=True)
         current_row += 1
         #получаем все вопросы из опроса
         questions = Question.objects.filter(poll_id = poll.id)
         ws.cell(row=current_row, column=1, value="ФИО")
         ws.cell(row=current_row, column=2, value="Филиал")
         cur_column = 3
         for question in questions:
            #заголовки таблицы c названиями вопросов
            ws.cell(row=current_row, column=cur_column, value=question.name)
            cur_column += 1
         current_row += 1
         for current_employee in current_employees:
            #заполняем фио и филиал
            ws.cell(row=current_row, column=1, value=current_employee.name)
            ws.cell(row=current_row, column=2, value=current_employee.filial.name)
            cur_column = 3
            for question in questions:
               #находим ответ на вопрос
               answer = Answer.objects.filter(question_id = question.id, login_id = current_employee.id).first()   
               answer_text = answer.name if answer else "-"
               ws.cell(row=current_row, column=cur_column, value=answer_text)
               cur_column += 1
            current_row += 1 
         current_row += 2 
               
   #сохранение отчета
      report_filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
      wb.save(report_filename)
      return report_filename
   
   except Exception as e:
      print(f"Ошибка при формировании отчета: {str(e)}")
      return None      
            
            
            